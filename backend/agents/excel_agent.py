"""
Excel Agent — generates a .xlsx workbook with 3 sheets using openpyxl.
No LLM call needed; pure data transformation.
"""
import json
import uuid
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ARTIFACTS_DIR = Path(__file__).parent.parent / "artifacts"


def _parse_json_list(raw) -> list:
    if not isinstance(raw, str):
        return raw if isinstance(raw, list) else []
    raw = raw.strip()
    start = raw.find("[")
    if start == -1:
        start = raw.find("{")
    if start == -1:
        return []
    result, _ = json.JSONDecoder().raw_decode(raw[start:])
    if isinstance(result, list):
        return result
    if isinstance(result, dict):
        return [result]
    if isinstance(result, str):
        return _parse_json_list(result)
    return []


HEADER_FILL = PatternFill(start_color="1A2E4A", end_color="1A2E4A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TEAL_FILL = PatternFill(start_color="008C8C", end_color="008C8C", fill_type="solid")


async def run_excel_agent(
    analysis_type: str,
    physician_list: str,
    dimensions: list[str] | None = None,
    icd10_codes: list[str] | None = None,
) -> dict:
    try:
        physicians = _parse_json_list(physician_list)

        wb = Workbook()

        _build_raw_sheet(wb, physicians)
        _build_pivot_sheet(wb, physicians)
        _build_icd10_sheet(wb, physicians, icd10_codes or [])

        # Remove default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        artifact_id = str(uuid.uuid4())
        out_path = ARTIFACTS_DIR / f"{artifact_id}.xlsx"
        wb.save(str(out_path))

        return {"artifact_id": artifact_id, "filename": "physician_analysis.xlsx"}

    except Exception as exc:
        return {"error": True, "message": str(exc)}


# ── Sheet builders ──────────────────────────────────────────────────────────

def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _build_raw_sheet(wb: Workbook, physicians: list):
    ws = wb.create_sheet("Physician Data")
    headers = [
        "ID", "NPI", "First Name", "Last Name", "Specialty",
        "Affiliation", "City", "State", "Volume Tier",
        "Total NSCLC Claims", "Board Certified",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for i, p in enumerate(physicians, start=2):
        row = [
            p.get("id", ""),
            p.get("npi", ""),
            p.get("firstName", ""),
            p.get("lastName", ""),
            p.get("specialty", ""),
            p.get("affiliation", ""),
            p.get("city", ""),
            p.get("state", ""),
            p.get("volumeTier", ""),
            p.get("totalNSCLCClaims", 0),
            "Yes" if p.get("boardCertified") else "No",
        ]
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = ALT_FILL

    _auto_width(ws)


def _build_pivot_sheet(wb: Workbook, physicians: list):
    ws = wb.create_sheet("Pivot State x Specialty")

    states = sorted({p.get("state", "Unknown") for p in physicians if p.get("state")})
    specialties = sorted({p.get("specialty", "Unknown") for p in physicians if p.get("specialty")})

    # Header row
    ws.cell(row=1, column=1, value="State \\ Specialty")
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).font = HEADER_FONT

    for col_idx, spec in enumerate(specialties, start=2):
        cell = ws.cell(row=1, column=col_idx, value=spec)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Aggregate claim volumes
    pivot: dict[tuple, int] = defaultdict(int)
    for p in physicians:
        pivot[(p.get("state", "Unknown"), p.get("specialty", "Unknown"))] += p.get("totalNSCLCClaims", 0)

    for row_idx, state in enumerate(states, start=2):
        ws.cell(row=row_idx, column=1, value=state).font = Font(bold=True)
        for col_idx, spec in enumerate(specialties, start=2):
            val = pivot.get((state, spec), 0)
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val else "")
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    _auto_width(ws)


def _build_icd10_sheet(wb: Workbook, physicians: list, icd10_codes: list[str]):
    ws = wb.create_sheet("ICD-10 Breakdown")

    all_codes = sorted(
        {code for p in physicians for code in p.get("icd10ClaimVolume", {}).keys()}
    )
    if icd10_codes:
        all_codes = [c for c in all_codes if c in icd10_codes] or all_codes

    headers = ["ICD-10 Code", "Physician Count", "Total Claims", "Avg Claims / Physician"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for i, code in enumerate(all_codes, start=2):
        matching = [p for p in physicians if code in p.get("icd10ClaimVolume", {})]
        total_claims = sum(p["icd10ClaimVolume"][code] for p in matching)
        avg = round(total_claims / len(matching), 1) if matching else 0
        row = [code, len(matching), total_claims, avg]
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = ALT_FILL

    _auto_width(ws)
